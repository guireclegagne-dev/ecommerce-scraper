import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DataExporter:
    """
    Gestionnaire d'export de données dans différents formats
    """
    
    def __init__(self, output_dir: str = "data/exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True, parents=True)
    
    def export_to_csv(self, data: List[Dict], filename: str = None, 
                     selected_fields: Optional[List[str]] = None) -> str:
        """
        Exporter les données en CSV
        """
        if not data:
            logger.warning("Aucune donnée à exporter")
            return None
        
        try:
            # Créer un DataFrame
            df = pd.DataFrame(data)
            
            # Sélectionner les champs si spécifié
            if selected_fields:
                # Mapper les noms français vers les noms de colonnes
                field_mapping = {
                    'Marque': 'marque',
                    'Modèle': 'modele',
                    'Finitions': 'finitions',
                    'Caractéristiques': 'caracteristiques',
                    'Prix': 'prix',
                    'Stock': 'disponibilite',
                    'URL': 'url',
                    'Date de collecte': 'date_collecte'
                }
                
                columns_to_keep = [field_mapping.get(field, field.lower()) 
                                  for field in selected_fields 
                                  if field_mapping.get(field, field.lower()) in df.columns]
                
                if columns_to_keep:
                    df = df[columns_to_keep]
            
            # Générer le nom de fichier
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            filepath = self.output_dir / filename
            
            # Exporter
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            logger.info(f"Export CSV réussi: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export CSV: {str(e)}")
            return None
    
    def export_to_excel(self, data: List[Dict], filename: str = None,
                       selected_fields: Optional[List[str]] = None,
                       sheet_name: str = "Produits") -> str:
        """
        Exporter les données en Excel avec mise en forme
        """
        if not data:
            logger.warning("Aucune donnée à exporter")
            return None
        
        try:
            # Créer un DataFrame
            df = pd.DataFrame(data)
            
            # Sélectionner les champs si spécifié
            if selected_fields:
                field_mapping = {
                    'Marque': 'marque',
                    'Modèle': 'modele',
                    'Finitions': 'finitions',
                    'Caractéristiques': 'caracteristiques',
                    'Prix': 'prix',
                    'Stock': 'disponibilite',
                    'URL': 'url',
                    'Date de collecte': 'date_collecte'
                }
                
                columns_to_keep = [field_mapping.get(field, field.lower()) 
                                  for field in selected_fields 
                                  if field_mapping.get(field, field.lower()) in df.columns]
                
                if columns_to_keep:
                    df = df[columns_to_keep]
            
            # Renommer les colonnes en français
            column_names_fr = {
                'marque': 'Marque',
                'modele': 'Modèle',
                'finitions': 'Finitions',
                'caracteristiques': 'Caractéristiques',
                'prix': 'Prix',
                'url': 'URL',
                'image': 'Image',
                'disponibilite': 'Disponibilité',
                'site_source': 'Site Source',
                'date_collecte': 'Date de collecte'
            }
            
            df = df.rename(columns=column_names_fr)
            
            # Générer le nom de fichier
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = self.output_dir / filename
            
            # Créer le writer Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Obtenir la feuille et formater
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Ajuster la largeur des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formater l'en-tête
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            logger.info(f"Export Excel réussi: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export Excel: {str(e)}")
            return None
    
    def export_to_json(self, data: List[Dict], filename: str = None,
                      selected_fields: Optional[List[str]] = None) -> str:
        """
        Exporter les données en JSON
        """
        if not data:
            logger.warning("Aucune donnée à exporter")
            return None
        
        try:
            import json
            
            # Filtrer les champs si nécessaire
            if selected_fields:
                field_mapping = {
                    'Marque': 'marque',
                    'Modèle': 'modele',
                    'Finitions': 'finitions',
                    'Caractéristiques': 'caracteristiques',
                    'Prix': 'prix',
                    'Stock': 'disponibilite',
                    'URL': 'url',
                    'Date de collecte': 'date_collecte'
                }
                
                filtered_data = []
                for item in data:
                    filtered_item = {
                        field_mapping.get(field, field.lower()): item.get(field_mapping.get(field, field.lower()))
                        for field in selected_fields
                        if field_mapping.get(field, field.lower()) in item
                    }
                    filtered_data.append(filtered_item)
                
                data = filtered_data
            
            # Générer le nom de fichier
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            filepath = self.output_dir / filename
            
            # Exporter
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Export JSON réussi: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export JSON: {str(e)}")
            return None
    
    def get_statistics(self, data: List[Dict]) -> Dict:
        """
        Générer des statistiques sur les données
        """
        if not data:
            return {}
        
        df = pd.DataFrame(data)
        
        stats = {
            'total_produits': len(df),
            'marques_uniques': df['marque'].nunique() if 'marque' in df.columns else 0,
            'sites_sources': df['site_source'].nunique() if 'site_source' in df.columns else 0,
        }
        
        # Produits par marque
        if 'marque' in df.columns:
            stats['produits_par_marque'] = df['marque'].value_counts().head(10).to_dict()
        
        # Produits par site
        if 'site_source' in df.columns:
            stats['produits_par_site'] = df['site_source'].value_counts().to_dict()
        
        # Disponibilité
        if 'disponibilite' in df.columns:
            stats['disponibilite'] = df['disponibilite'].value_counts().to_dict()
        
        return stats
    
    def create_summary_report(self, data: List[Dict], filename: str = None) -> str:
        """
        Créer un rapport récapitulatif en Excel avec statistiques
        """
        if not data:
            logger.warning("Aucune donnée pour le rapport")
            return None
        
        try:
            # Générer le nom de fichier
            if not filename:
                filename = f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = self.output_dir / filename
            
            # Créer le DataFrame principal
            df = pd.DataFrame(data)
            
            # Renommer les colonnes
            column_names_fr = {
                'marque': 'Marque',
                'modele': 'Modèle',
                'finitions': 'Finitions',
                'caracteristiques': 'Caractéristiques',
                'prix': 'Prix',
                'disponibilite': 'Disponibilité',
                'site_source': 'Site Source',
                'date_collecte': 'Date de collecte'
            }
            df = df.rename(columns=column_names_fr)
            
            # Créer le writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Feuille principale
                df.to_excel(writer, sheet_name='Données', index=False)
                
                # Feuille statistiques
                stats = self.get_statistics(data)
                
                stats_data = []
                stats_data.append(['Statistiques Générales', ''])
                stats_data.append(['Total de produits', stats.get('total_produits', 0)])
                stats_data.append(['Marques uniques', stats.get('marques_uniques', 0)])
                stats_data.append(['Sites sources', stats.get('sites_sources', 0)])
                stats_data.append(['', ''])
                
                if 'produits_par_marque' in stats:
                    stats_data.append(['Produits par Marque', ''])
                    for marque, count in stats['produits_par_marque'].items():
                        stats_data.append([marque, count])
                    stats_data.append(['', ''])
                
                if 'produits_par_site' in stats:
                    stats_data.append(['Produits par Site', ''])
                    for site, count in stats['produits_par_site'].items():
                        stats_data.append([site, count])
                
                stats_df = pd.DataFrame(stats_data, columns=['Indicateur', 'Valeur'])
                stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
                
                # Formater les feuilles
                from openpyxl.styles import Font, PatternFill, Alignment
                
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Ajuster les colonnes
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formater l'en-tête
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            logger.info(f"Rapport créé avec succès: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur lors de la création du rapport: {str(e)}")
            return None